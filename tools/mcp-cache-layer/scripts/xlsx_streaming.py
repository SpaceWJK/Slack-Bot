# -*- coding: utf-8 -*-
"""
xlsx_streaming.py — XLSX 파일 직접 스트리밍 파서 (openpyxl read_only)

GDI MCP 텍스트 청크 경로(reconstructors.py)와 독립적으로,
원본 .xlsx 파일을 직접 읽어야 하는 경우에 사용한다.

openpyxl read_only=True + iter_rows(values_only=True) 스트리밍으로
대용량 파일(3000행+)을 메모리 효율적으로 파싱한다.

공개 API:
  - stream_xlsx(path, sheet_name=None, max_rows=20000) -> dict
  - get_sheet_names(path) -> list[str]
  - parse_xlsx_to_markdown(path, sheet_name=None, max_rows=20000) -> str
"""

import tracemalloc
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    raise ImportError(
        "xlsx_streaming requires openpyxl>=3.1.0. "
        "Install with: pip install openpyxl"
    ) from e


def _iter_sheet(ws, max_rows: int):
    """시트에서 헤더와 데이터 행을 lazy 순회하여 반환한다.

    Args:
        ws: openpyxl read_only Worksheet
        max_rows: 헤더 제외 데이터 행 수 상한 (>= 1)

    Returns:
        (headers: list[str], rows: list[list[str]], truncated: bool)
    """
    row_iter = ws.iter_rows(values_only=True)

    first = next(row_iter, None)
    if first is None:
        return [], [], False
    headers = [str(c) if c is not None else "" for c in first]

    rows = []
    truncated = False
    for i, row in enumerate(row_iter):
        if i >= max_rows:
            truncated = True
            break
        rows.append([str(c) if c is not None else "" for c in row])

    return headers, rows, truncated


def stream_xlsx(path, sheet_name: str = None, max_rows: int = 20000) -> dict:
    """XLSX 파일을 read_only 스트리밍으로 파싱하여 구조화된 dict를 반환한다.

    Args:
        path: XLSX 파일 경로 (str 또는 Path)
        sheet_name: 처리할 시트명. None이면 전체 시트 처리.
        max_rows: 헤더 제외 데이터 행 수 상한. 기본 20000.

    Returns:
        {
            "sheets": [
                {
                    "name": str,
                    "headers": list[str],
                    "rows": list[list[str]],
                    "row_count": int,   # 반환된 데이터 행 수 (헤더 제외)
                    "truncated": bool,
                }
            ],
            "total_sheets": int,
            "path": str,
            "_memory_peak_mb": float,  # tracemalloc peak (MB)
        }

    Raises:
        ValueError: max_rows < 1
        FileNotFoundError: 파일 없음
        openpyxl.utils.exceptions.InvalidFileException: 손상된 XLSX
    """
    if max_rows < 1:
        raise ValueError(f"max_rows must be >= 1, got {max_rows}")

    path = str(path)
    if not Path(path).exists():
        raise FileNotFoundError(f"XLSX file not found: {path}")

    tracemalloc.start()
    wb = None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        target_sheets = wb.sheetnames
        if sheet_name is not None:
            if sheet_name not in target_sheets:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. "
                    f"Available: {target_sheets}"
                )
            target_sheets = [sheet_name]

        sheets_result = []
        for sname in target_sheets:
            ws = wb[sname]
            headers, rows, truncated = _iter_sheet(ws, max_rows)
            sheets_result.append({
                "name": sname,
                "headers": headers,
                "rows": rows,
                "row_count": len(rows),
                "truncated": truncated,
            })

    finally:
        if wb is not None:
            wb.close()
        _, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()

    return {
        "sheets": sheets_result,
        "total_sheets": len(sheets_result),
        "path": path,
        "_memory_peak_mb": round(peak / 1024 / 1024, 2),
    }


def get_sheet_names(path) -> list:
    """XLSX 파일의 시트명 목록을 반환한다.

    Args:
        path: XLSX 파일 경로

    Returns:
        list[str] — 시트명 목록
    """
    path = str(path)
    if not Path(path).exists():
        raise FileNotFoundError(f"XLSX file not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        names = list(wb.sheetnames)
    finally:
        wb.close()
    return names


def parse_xlsx_to_markdown(path, sheet_name: str = None, max_rows: int = 20000) -> str:
    """XLSX 파일을 스트리밍으로 파싱하여 마크다운 테이블 문자열을 반환한다.

    Args:
        path: XLSX 파일 경로
        sheet_name: 처리할 시트명. None이면 전체 시트.
        max_rows: 데이터 행 수 상한.

    Returns:
        str — 마크다운 형식 테이블 (시트별 ## 헤더 포함)
    """
    result = stream_xlsx(path, sheet_name=sheet_name, max_rows=max_rows)
    lines = []

    for sheet in result["sheets"]:
        lines.append(f"## {sheet['name']}")
        if not sheet["headers"]:
            lines.append("_(빈 시트)_")
            lines.append("")
            continue

        lines.append("| " + " | ".join(sheet["headers"]) + " |")
        lines.append("|" + "|".join("---" for _ in sheet["headers"]) + "|")

        for row in sheet["rows"]:
            lines.append("| " + " | ".join(row) + " |")

        if sheet["truncated"]:
            lines.append(f"\n_(… max_rows={max_rows} 초과 — 이후 행 생략)_")

        lines.append("")

    return "\n".join(lines)
